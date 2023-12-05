import time
import logging
import openpyxl
import aiohttp
import asyncio
import requests
import os
import uuid
import botocore
import operator
import re
import datetime
import urllib.parse

from urllib.parse import quote, quote_plus

from src.config.config import baseUrl,message_str

logger = logging.getLogger(__name__)

def tab_log(tab):
    return "└── [" + tab + "]"
def timer(function):
    def wrapper(*args, **kws):
        t_start = time.time()
        result = function(*args, **kws)
        t_end = time.time()
        t_count = t_end - t_start
        logger.info(
            f"<function {function.__qualname__}> - Time Coast: {t_count:.2f}s \n"
        )
        return result

    return wrapper


def string_to_timestamp(time_string):
    if time_string.isdigit():
        return int(time_string)
    date = datetime.datetime.strptime(time_string, "%Y-%m-%d %H:%M:%S")  # Thay đổi định dạng ngày tháng nếu cần
    return int(date.timestamp() * 1000)  # Chuyển đổi thành timestamp và nhân với 1000 để có dạng miligisecond

def get_value_from_query(query, field):
    # Tách các cặp key-value trong query bằng phương thức split('&')
    pairs = query.split('&')

    # Kiểm tra và loại bỏ các phần tử rỗng sau khi phân tách query
    pairs = [pair for pair in pairs if pair]

    # Tìm cặp key-value chứa trường field và trả về giá trị của trường đó
    for pair in pairs:
        key_value = pair.split('=', 1)  # Phân tách thành tối đa một cặp
        if len(key_value) == 2 and key_value[0] == field:
            return key_value[1]

    # Nếu không tìm thấy trường field trong query, trả về None
    return None



def read_excel_to_array(file_path,is_filter = True):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    data_array = []

    for row in sheet.iter_rows(values_only=True):
        if row:
            if is_filter:
                data_array.append(list(filter(filter_not_null_and_not_header_row, row)))
            else:
                data_array.append(row)
    wb.close()

    return data_array


def filter_not_null_and_not_header_row(i):
    headers = ["koc", "hashtags", "url", "channel","keyword","Keyword/hashtag","Start time","End time","HASHTAGS","KOC"]
    if isinstance(i, (int, float)) and i == 0 and i not in headers and i :
        return True

    return i

def read_excel_to_array_only_1_row(file_path,is_filter = True,column_index=1):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    data_array = []

    for row in sheet.iter_rows(values_only=True):
        if row:
            if is_filter:
                filtered_row = list(filter(filter_not_null_and_not_header_row, row))
                if filtered_row:
                    data_array.append([filtered_row[column_index - 1]])
            else:
                data_array.append([row[column_index - 1]])
    wb.close()

    return data_array

def get_query(p):
    str = "?"
    for key, value in p.items():
        str += f"&{key}={value}"
    return (str)

def get_query_in_array(arr):
    arr_strings = [remove_special_characters(str(i)) for i in arr if isinstance(i, str)]
    return "&" + "&".join(["keywords=" + quote(i.encode()) for i in arr_strings])


async def update_progress_input_file(body):
    url = baseUrl + "input-file/" + body['id'] + "/update-progress"
    async with aiohttp.ClientSession() as session:
        async with session.put(url + get_query(body)) as response:
            if response.status == 200:
                data = await response.json()
                return data.get('data',data)
            else:
                return False

async def get_input_file_status(id):
    url = f"{baseUrl}input-file/{id}"
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            if response.status == 200:
                data = await response.json()
                return data.get('data', {}).get('status','')
            else:
                return False

#
# async def update_progress_input_file(data):
#     url = baseUrl + "input-file/" + data['id'] +"/update-progress"
#     response = requests.put(url, data)
#     print(response)
def group_by_column(data, column_index):
    grouped_data = {}

    for row in data:
        key = row[column_index]
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(row)
    return grouped_data
def group_by_field(data, field, order_by_field="created_at", ascending=True):
    grouped_data = {}

    for item in data:
        key = item[field]
        if key not in grouped_data:
            grouped_data[key] = []
        grouped_data[key].append(item)

    if order_by_field:
        # Sort the grouped data by the specified field
        for key, group in grouped_data.items():
            grouped_data[key] = sorted(group, key=operator.itemgetter(order_by_field), reverse=not ascending)

    return grouped_data


def remove_excel_file(file_path):
    is_excel_path = file_path.split(".")[1] == "xlsx"
    if is_excel_path and os.path.exists(file_path):
        os.remove(file_path)
        print(f"Delete successfully file: {file_path}")
    else:
        print(f"File {file_path} does not exist.")

async def download_and_read_excel_file(url,path,only_1_row=True):
    async def wait_for_aiohttp():
        await asyncio.sleep(5)  # Chờ trong 5 giây trước khi tiếp tục với aiohttp

    # Đợi trong 5 giây trước khi tiếp tục với aiohttp
    await wait_for_aiohttp()

    async with aiohttp.ClientSession() as session:
        async with session.get(url) as response:
            if response.status == 200:
                random_uuid = uuid.uuid4()
                filename = f"{path}-{random_uuid}.xlsx"
                with open(filename, "wb") as f:
                    while True:
                        chunk = await response.content.read(1024)
                        if not chunk:
                            break
                        f.write(chunk)
                if only_1_row == True:
                    arr = read_excel_to_array_only_1_row(filename)
                else:
                    arr = read_excel_to_array(filename,is_filter=only_1_row)

                remove_excel_file(filename)
                return arr

def init_chunks(rows, index_per_chunk=10):
    if index_per_chunk <= 0:
        raise ValueError("Chunk size must be greater than 0")

    chunks = []
    chunk = []

    for row in rows:
        for keyword in row:
            if len(chunk) < index_per_chunk:
                chunk.append(keyword)
            else:
                chunks.append(chunk.copy())
                chunk.clear()
                chunk.append(keyword)

    if chunk:
        chunks.append(chunk)

    return chunks


def init_chunks_v_array(rows, index_per_chunk=10):
    if index_per_chunk <= 0:
        raise ValueError("Chunk size must be greater than 0")

    chunks = []
    chunk = []

    for row in rows:
        if len(chunk) < index_per_chunk:
            chunk.append(row)
        else:
            chunks.append(chunk.copy())
            chunk.clear()
            chunk.append(row)

    if chunk:
        chunks.append(chunk)

    return chunks


async def fetch_data(session, url):
    async with session.get(url) as response:
        if response.status == 200:
            data = await response.json()
            return data['data']
        else:
            return None

async def get_queue_of_tab():
    url = baseUrl + "input-file/queue/"
    print(url)
    async with aiohttp.ClientSession() as session:
        response = await fetch_data(session, url)
        return response

def array_find_field(arr,field,compare):
    return next((i for i in arr if i[field] == compare), None)

def array_filter_field(arr,field,compare):
    return [i for i in arr if i.get(field,None) == compare]

def get_platform(link):
    if 'youtube.com' in link or 'youtu.be' in link:
        return 'youtube'
    if 'tiktok.com' in link:
        return 'tiktok'
    if 'fb.com' in link or 'facebook.com' in link:
        return 'facebook'
    if 'instagram.com' in link:
        return 'instagram'


def get_post_id_by_platform(url):
    # Kiểm tra từng loại URL để trích xuất id
    # TikTok URL
    tiktok_match = re.search(r'video/(\d+)', url)
    if tiktok_match:
        return tiktok_match.group(1)

    # YouTube URL
    youtube_match = re.search(r'watch\?v=([a-zA-Z0-9_-]+)', url)
    if youtube_match:
        return youtube_match.group(1)

    # Instagram URL
    instagram_match = re.search(r'\/p\/([a-zA-Z0-9_-]+)\/', url)
    if instagram_match:
        return instagram_match.group(1)

    return None


def remove_illegal_characters(text):
    # Chuyển đổi cell_value thành string nếu nó là kiểu int hoặc float
    try:
        if isinstance(text, (int, float, type(None))):
            text = str(text)

        # Các ký tự không hợp lệ trong tên ô Excel (bổ sung \x03 vào danh sách)
        illegal_characters = r'[\x00-\x1F\\*?\[\]]'
        return re.sub(illegal_characters, '', text)
    except Exception as e:
        logger.info(f"Cannot remove_illegal_characters: {e}")
        return text

def add_row_to_excel(sheet,row):
    cleaned_row = [remove_illegal_characters(cell_value) for cell_value in row]
    sheet.append(cleaned_row)
    return sheet

def is_valid_timestamp(timestamp):
    try:
        datetime.datetime.fromisoformat(timestamp)
        return True
    except ValueError:
        return False

def timestamp_to_date_string(timestamp):
    if timestamp is None:
        return ""

    # Kiểm tra và chuyển đổi timestamp thành số nguyên nếu nó là một chuỗi
    if isinstance(timestamp, str):
        dt_object = datetime.datetime.fromisoformat(timestamp)
    else:
        dt_object = datetime.datetime.fromtimestamp(timestamp)

    # Chuyển đổi đối tượng datetime thành chuỗi ngày tháng
    date_string = dt_object.strftime('%Y-%m-%d %H:%M:%S')  # Định dạng có thể thay đổi tùy ý

    return date_string

def is_valid_url(url, platform_url):
    return is_valid_http_url(url) and platform_url in url

def is_valid_http_url(url):
    regex = r"^(http|https):\/\/"
    return re.match(regex, url)

def get_tiktok_username(string):
    try:
        if string:
            string = str(string).strip()
            platform_url = 'tiktok.com'
            if is_valid_url(string, platform_url):
                match = re.search(r'@(.+?)(?:\?|$)', string)
                if match:
                    return match.group(1)
            if '@' in string:
                return string.replace('@', '')
        print("username1",string)
        if '/' in string:
            string = username.replace('/', '')
        print("username2",string)
        return string
    except Exception as e:
        print(f"An error occurred: {e}")
        return string



def channel_handle_unique_koc_sheet_data(file_path, sheet_koc_name):
    wb = openpyxl.load_workbook(file_path)
    sheet_kocs = wb[sheet_koc_name]

    # Đọc dữ liệu cũ từ tệp Excel
    posts_data = read_excel_to_array(file_path,False)
    print("posts_data",posts_data)
    posts_data.pop(0)  # Loại bỏ tiêu đề nếu có

    # Nhóm dữ liệu theo cột 2 (koc)
    posts_data_group_koc = group_by_column(posts_data, 2)

    for koc, koc_data_list in posts_data_group_koc.items():
        first_data = koc_data_list[0]
        total_posts = len(koc_data_list)
        followers = first_data[9] if len(first_data) > 9 else 0
        views = 0
        likes = 0
        related_contents = ""
        latest_contents = first_data[8] if len(first_data) > 8 else 0

        for koc_data in koc_data_list:
            related_contents += (koc_data[0] + ",\n")
            views += int(koc_data[3]) if koc_data[3] is not None and koc_data[3].isdigit() else 0
            likes += int(koc_data[4]) if koc_data[4] is not None and koc_data[4].isdigit() else 0
            if latest_contents < koc_data[8]:
                latest_contents = koc_data[8]

        row_koc = [koc, related_contents, total_posts, followers, views, likes, latest_contents]
        sheet_kocs.append(row_koc)

    # Lưu và đóng tệp
    wb.save(file_path)
    wb.close()

    return file_path

def url_encode_string(input_string):
    # Sử dụng hàm quote() từ urllib.parse để encode chuỗi thành dạng URL
    encoded_string = urllib.parse.quote(input_string)
    return encoded_string


def remove_special_characters(input_string):
    try:
        # Sử dụng biểu thức chính quy (regex) để loại bỏ tất cả các ký tự đặc biệt trong chuỗi đầu vào
        cleaned_string = re.sub(r'[^\w\s]', '', input_string)
    except TypeError:
        # Xử lý trường hợp không phải là chuỗi hoặc là ký tự rỗng hoặc số
        cleaned_string = input_string

    return cleaned_string

def convert_to_date_format(time_str):
    year = time_str[:4]
    month = time_str[4:6]
    day = time_str[6:8]

    return f"{year}-{month}-{day}"


def extract_result_from_excel_formula(formula):
    try:
        parts = formula.split('),"')
        if len(parts) > 1:
            result = parts[1].strip(')"')
            return result
        else:
            return formula
    except Exception as e:
        return formula


def count_word_occurrences(word, s):
    try:
        regex = re.compile(word)
        matches = regex.findall(s)

        return len(matches)
    except Exception as e:
        return 0

def set_file_path(tab):
    timestamp_now = int(time.time())
    file_name = f"data-results-{tab}-{timestamp_now}.xlsx"
    print("File name init: ",file_name)
    return file_name


