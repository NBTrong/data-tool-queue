import time
import logging
import openpyxl
import aiohttp
import asyncio
import requests
import os
import uuid
import botocore
import operator
from concurrent.futures import ThreadPoolExecutor
import datetime  # Add this line to import the datetime module

from urllib.parse import quote, quote_plus

from src.config.config import baseUrl, message_str
from src.utils.api.apis import import_rows_api
from src.utils.time_helper import convert_timestamp_to_date_string
from src.utils.utils import tab_log, add_row_to_excel, timestamp_to_date_string, read_excel_to_array, group_by_column, \
    update_progress_input_file, init_chunks, channel_handle_unique_koc_sheet_data, remove_illegal_characters, \
    get_value_from_query, download_and_read_excel_file, set_file_path

logger = logging.getLogger(__name__)

tab = "keyword-explore"

keyword_explore_post = ["Video", "Description", "KOC", "Views", "Likes", "Comments", "Shares", "Save", "Uploaded",
                        "KOC follower"]
keyword_explore_kocs = ["Author", "Related Contents", "Videos", "Followers", "Views", "Likes", "Latest Related Content"]


def data_failed_default(keyword):
    return {
        'keyword': keyword,
        'name': f"This {keyword} could not be found any post",
        'title': f"This {keyword} could not be found any post",
        'desc': f"This {keyword} could not be found any post",
    }


def tiktok_data_parse(data, keyword, is_hashtag, input_file_id, filters=None):
    if is_hashtag:
        posts = []
        print("-----------data",data)
        if data and 'aweme_list' in data:
            for item in data['aweme_list']:
                post = {
                    "keyword": keyword,
                    "platform": "tiktok",
                    'input_file_id': input_file_id,
                    'description': item.get('desc'),
                    'thumb_url': item['video'].get('thumb'),
                    'post_url':  item['video'].get('url'),
                    'koc': item.get('author')['unique_id'],
                    'total_comments': item.get('total_comments'),
                    'total_likes': item.get('total_likes'),
                    'total_saves': item.get('total_saves'),
                    'total_shares': item.get('total_shares'),
                    'total_views': item.get('total_views'),
                    'uploaded_time': convert_timestamp_to_date_string(item.get('created_at')),
                }
                posts.append(post)
    else:
        posts = []
        if data is not None and 'aweme_list' in data:
            for item in data['aweme_list']:
                post = {
                    "keyword": keyword,
                    "platform": "tiktok",
                    'input_file_id': input_file_id,
                    'description': item.get('desc'),
                    'thumb_url': item['video'].get('thumb'),
                    'post_url':  item['video'].get('url'),
                    'koc': item.get('author')['unique_id'],
                    'total_comments': item.get('total_comments'),
                    'total_likes': item.get('total_likes'),
                    'total_saves': item.get('total_saves'),
                    'total_shares': item.get('total_shares'),
                    'total_views': item.get('total_views'),
                    'uploaded_time': convert_timestamp_to_date_string(item.get('created_at')),
                }
                posts.append(post)

    if filters:
        posts = filter_data(posts, filters)

    return posts


def filter_data(data, filters):
    if not filters:
        return data

    from_date = filters.get("from", "")
    to_date = filters.get("to", "")
    result = data
    print("from_date", from_date)
    print("to_date", to_date)

    if from_date:
        result = list(filter(lambda item: item['created_at'] >= from_date, result))
    if to_date:
        result = list(filter(lambda item: item['created_at'] <= to_date, result))

    return result


def fetch_tiktok_keyword_explore_hashtag_id_api(keyword):
    print("keyword", keyword)
    if keyword.find("#") != -1:
        search = keyword.replace("#", "")
        api_url = baseUrl + f"crawl/tiktok/hashtag-id?hashtag={search}"
        max_retry = 5  # Số lần tối đa gọi lại API khi thất bại
        retry_count = 0

        while retry_count < max_retry:
            print(api_url)
            response = requests.get(api_url)
            if response.status_code == 200:
                res_data = response.json()
                hid = res_data['data']
                if str(hid).isdigit():
                    return hid
                else:
                    retry_count += 1
            else:
                mess_err = tab_log(tab) + f": Retry to get hashtag id: {api_url}"
                logger.info(mess_err)
                retry_count += 1

        logger.info(tab_log(tab) + f": Can not find hashtag: {search}")
        # Đã gọi API quá 3 lần mà vẫn thất bại
        return keyword
    else:
        return keyword


def fetch_tiktok_keyword_explore_api(keyword, is_hashtag, publish_time, input_file_id,search):
    results = []

    if (not is_hashtag and keyword) or (is_hashtag and keyword.isdigit()):
        base_api_url = baseUrl + f"crawl/tiktok/search/post?keyword={keyword}&publish_time={publish_time}"
        if is_hashtag:
            base_api_url = baseUrl + f"crawl/tiktok/hashtag-posts?cid={keyword}"
        max_retry = 3  # Số lần tối đa gọi lại API khi thất bại
        retry_count = 0
        has_more = 1
        offset = 0

        while has_more and retry_count < max_retry:
            if is_hashtag:
                api_url = base_api_url + f"&cursor={offset}"
            else:
                api_url = base_api_url + f"&offset={offset}"
            print(api_url)
            response = requests.get(api_url)
            if response.status_code == 200:
                res_data = response.json()
                data = res_data['data']
                offset = data.get('cursor', None)
                if offset:
                    retry_count = 0
                    has_more = data.get("has_more",0) and len(data.get("aweme_list"))
                    if is_hashtag:
                        new_data = tiktok_data_parse(data, search, is_hashtag, input_file_id)
                    else:
                        new_data = tiktok_data_parse(data, keyword, is_hashtag, input_file_id)
                    if len(new_data):
                        results.append(new_data)
                else:
                    retry_count += 1

            else:
                print(tab_log(tab) + f": Retry to call api {tab}")
                retry_count += 1

    return results
    # return data_failed_default(keyword)


def keyword_explore_get_row(item):
    video_url = item['video']['url'] if 'video' in item and 'url' in item['video'] else None
    video_name = item['name'] if 'name' in item else None
    koc = item['author']['unique_id'] if 'author' in item and 'unique_id' in item['author'] else None
    followers = item['author']['follower_count'] if 'author' in item and 'follower_count' in item['author'] else None
    views = item['total_views'] if 'total_views' in item else None
    likes = item['total_likes'] if 'total_likes' in item else None
    comments = item['total_comments'] if 'total_comments' in item else None
    shares = item['total_shares'] if 'total_shares' in item else None
    save = item['total_saves'] if 'total_saves' in item else None
    uploaded = item['created_at'] if 'created_at' in item else None

    return [video_url, video_name, koc, views, likes, comments, shares, save, timestamp_to_date_string(uploaded),
            followers]


def parse_data_kocs(data):
    result = {}

    for video in data:
        if video and len(video):
            author_id = video.get('author', {}).get('unique_id')

            if author_id not in result:
                result[author_id] = {
                    'author': video.get('author'),
                    'total_views': 0,
                    'total_likes': 0,
                    'videos': [],
                    'latest_content': None,
                }

            result[author_id]['total_views'] += video.get('total_views', 0)
            result[author_id]['total_likes'] += video.get('total_likes', 0)

            if result[author_id]['latest_content'] is None or video.get('created_at') > result[author_id][
                'latest_content']:
                result[author_id]['latest_content'] = video.get('created_at')

            result[author_id]['videos'].append(video.get('video'))

    return list(result.values())


def gen_keyword_vn(keyword):
    if "vn" not in keyword:
        keyword = f"{keyword} vn"
    return keyword


def tiktok_keyword_explore_process_chunk(chunk, publish_time, input_file_id, max_workers=3):
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # get id and platform from url in same time
        keywords_parse = [(fetch_tiktok_keyword_explore_hashtag_id_api(keyword), keyword.find("#") != -1) for keyword in
                          chunk]

        # keywords_vn = [(gen_keyword_vn(keyword), keyword.find("#") != -1) for keyword in
        #                chunk if keyword.find("#") == -1]
        #
        # merge_keywords = keywords_parse + keywords_vn
        #
        # print("keywords_parse", keywords_parse)
        # print("keywords_vn", keywords_vn)
        # print("merge_keywords", merge_keywords)
        # print("ids_and_platforms",ids_and_platforms)
        # call api
        results = list(
            executor.map(
                lambda idx_keyword: fetch_tiktok_keyword_explore_api(idx_keyword[1][0], idx_keyword[1][1], publish_time,
                                                                     input_file_id, chunk[0]),
                enumerate(keywords_parse))
        )

    return results


async def tiktok_keyword_explore_crawl_data(rows, input_file_id, query, tab="keyword-explores"):
    max_workers = 2

    # Initialize chunks
    chunks = init_chunks(rows, max_workers)

    try:
        publish_time = get_value_from_query(query, "publish_time")
        if not publish_time:
            publish_time = 0

        print("publish_time", publish_time)
        # Loop through chunks to crawl data
        old_process = 0
        for index, chunk in enumerate(chunks):
            try:
                results = tiktok_keyword_explore_process_chunk(chunk, publish_time, input_file_id, max_workers)

                for result in results:
                    for rows in result:
                        #this will import default 30 row
                        print('rows',rows)
                        import_rows_api(rows, tab)

                # update process
                now_process = int(((index + 1) / len(chunks)) * 100)
                if now_process - old_process > 0:
                    old_process = now_process
                    update_data = await update_progress_input_file({'id': input_file_id, 'progress': now_process})
                    # handle stop
                    if update_data.get('status', None) == "stop":
                        logger.info(f"{tab_log(tab)}: Stop process")
                        break


            except Exception as e:
                logger.info(f"{tab_log(tab)}: chunk: {chunk} | error: {e}")

    except Exception as e:
        # In case of any exception, save the workbook and close it before re-raising the exception
        raise e
    print("tab",tab)
    return tab



# ---------------------------------------------Youtube---------------------------------------------

def youtube_keyword_explore_process_chunk(chunk, max_workers=3):
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        results = list(
            executor.map(lambda keyword: fetch_youtube_keyword_explore_api(keyword), chunk))

    return results


async def youtube_keyword_explore_crawl_data(rows, input_file_id, tab="youtube-keyword-explore"):
    # Initialize variables
    file_path = set_file_path(tab)
    max_workers = 2
    sheet_koc_name = "Unique KOCs"

    wb = openpyxl.Workbook()
    # Select the active sheet (you can modify this depending on your Excel structure)
    sheet_posts = wb.create_sheet(title="Posts")
    sheet_kocs = wb.create_sheet(title=sheet_koc_name)

    # Initialize header rows
    sheet_posts.append(keyword_explore_post)
    sheet_kocs.append(keyword_explore_kocs)

    # Remove default sheet
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    # Save the workbook once before processing chunks
    wb.save(file_path)

    # Initialize chunks
    chunks = init_chunks(rows, max_workers)

    try:
        # Loop through chunks to crawl data
        old_process = 0
        for index, chunk in enumerate(chunks):
            try:
                results = youtube_keyword_explore_process_chunk(chunk, max_workers)

                # update process
                now_process = int(((index + 1) / len(chunks)) * 100)
                if now_process - old_process > 0:
                    old_process = now_process
                    update_data = await update_progress_input_file({'id': input_file_id, 'progress': now_process})
                    # handle stop
                    if update_data.get('status', None) == "stop":
                        logger.info(f"{tab_log(tab)}: Stop process")
                        break

                # Append results to the sheet
                for result in results:
                    for row in result:
                        try:
                            sheet_posts.append(row)
                        except Exception as e:
                            logger.info(f"{tab_log(tab)}: row: {row} | error: {e}")

            except Exception as e:
                wb.save(file_path)
                logger.info(f"{tab_log(tab)}: chunk: {chunk} | error: {e}")

        # Save the workbook after processing all chunks
        wb.save(file_path)

        # Close the workbook before processing KOC data
        wb.close()

        # Process the KOC data separately
        channel_handle_unique_koc_sheet_data(file_path, sheet_koc_name)

        return file_path

    except Exception as e:
        # In case of any exception, save the workbook and close it before re-raising the exception
        wb.save(file_path)
        wb.close()
        raise e


def valid_data_youtube(data):
    return len(data)


def fetch_youtube_keyword_explore_api(keyword):
    base_api_url = baseUrl + f"crawl/youtube/search/posts?keyword={keyword}&limit=10"
    max_retry = 3  # Số lần tối đa gọi lại API khi thất bại
    retry_count = 0
    has_more = 1
    page = 1
    max_page = 20

    results = []

    while has_more and retry_count < max_retry:
        api_url = base_api_url + f"&page={page}"
        response = requests.get(api_url)
        print(api_url)
        if response.status_code == 200:
            res_data = response.json()
            data = res_data['data']
            valid = valid_data_youtube(data)
            if valid:
                retry_count = 0
                new_data = youtube_parse_data(data, keyword, page)
                for item in new_data:
                    row_posts = keyword_explore_get_row(item)
                    cleaned_row = [remove_illegal_characters(cell_value) for cell_value in row_posts]
                    results.append(cleaned_row)
                if page < max_page:
                    page += 1
                    has_more = 1
                else:
                    has_more = 0
            else:
                has_more = 0
                retry_count += 1

        else:
            print(tab_log(tab) + f": Retry to call api {tab}")
            retry_count += 1

    return results
    # return data_failed_default(keyword)


def youtube_parse_data(data, keyword=None, page=None):
    videos = []
    for item in data:
        video = {
            "id": item.get("id"),
            "name": item.get("title"),
            "keyword": keyword,
            "page": page,
            "video": {
                "thumb": item.get("thumbnail", False),
                "url": f"https://www.youtube.com/watch?v={item.get('id')}",
            },
            "author": {
                "name": item.get("channel", {}).get("name"),
                "avatar_url": item.get("channel", {}).get("avatarThumb", False),
                "unique_id": item.get("channel", {}).get("id"),
                "follower_count": item.get("channel", {}).get("channel_follower_count"),
            },
            "total_comments": item.get("commentCount", 0),
            "total_likes": item.get("likeCount", 0),
            "total_views": item.get("viewCount", 0),
            "created_at": item.get("uploadDate"),
        }
        videos.append(video)

    return videos


def get_tiktok_filters(rows):
    # Kiểm tra xem danh sách có ít nhất một hàng và các cột "hashtags", "duration start", "duration end" hay không
    if len(rows) < 1 or not all(col in rows[0] for col in ["Start time", "End time"]):
        return None

    # Xác định vị trí của các cột
    start_date_index, end_date_index = (
        rows[0].index("Start time"),
        rows[0].index("End time"),
    )

    start_date = rows[1][start_date_index]
    end_date = rows[1][end_date_index]

    # Kiểm tra nếu start_date và end_date là chuỗi, chuyển chúng thành datetime objects
    if isinstance(start_date, str):
        start_date = datetime.datetime.strptime(start_date, "%Y/%m/%d")
    if isinstance(end_date, str):
        end_date = datetime.datetime.strptime(end_date, "%Y/%m/%d")

    start_timestamp = int(start_date.timestamp())
    end_timestamp = int(end_date.timestamp())

    return {
        "from": start_timestamp,
        "to": end_timestamp,
    }
