import time
import logging
import openpyxl
import aiohttp
import asyncio
import requests
import os
import uuid
import botocore
import operator
import datetime
from concurrent.futures import ThreadPoolExecutor
import pandas as pd
import re

from urllib.parse import quote, quote_plus

from src.config.config import baseUrl,  message_str
from src.utils.api.apis import import_rows_api
from src.utils.time_helper import convert_timestamp_to_date_string
from src.utils.utils import tab_log, add_row_to_excel, timestamp_to_date_string, read_excel_to_array, group_by_column, \
    get_tiktok_username, channel_handle_unique_koc_sheet_data, init_chunks, is_valid_http_url, \
    update_progress_input_file, remove_illegal_characters, convert_to_date_format, get_value_from_query, \
    download_and_read_excel_file, string_to_timestamp, init_chunks_v_array, set_file_path

logger = logging.getLogger(__name__)

tab = "channel-explores"

channel_explore_header_row_post = ["Video", "Description", "KOC", "Views", "Likes", "Comments", "Shares", "Save", "Uploaded",
                        "KOC follower"]
channel_explore_header_row_kocs = ["Author", "Related Contents", "Videos", "Followers", "Views", "Likes", "Latest Related Content"]


def data_failed_default(keyword):
    return [f"Could not find channel: {keyword}", f"Could not find channel: {keyword}", keyword, '0', '0', '0', '0',
            '0', '0', '0']


def instagram_failed_default(keyword):
    return {
        'koc': keyword,
        'name': f"This {keyword} could not be found any post",
        'title': f"This {keyword} could not be found any post",
        'desc': f"This {keyword} could not be found any post",
    }


def parse_user_data(data):
    return {
        "id": data.get("userInfo", {}).get("user", {}).get("id"),
        "videoCount": data.get("userInfo", {}).get("stats", {}).get("videoCount",0),
    }


def parse_data(data,keyword,input_file_id, filters):
    posts = []
    if data is not None and 'aweme_list' in data:
        for item in data['aweme_list']:
            post = {
                "platform": "tiktok",
                'input_file_id': input_file_id,
                'description': item.get('desc'),
                'thumb_url': item['video'].get('thumb'),
                'post_url': item['video'].get('url'),
                'koc': item.get('author')['unique_id'],
                'total_comments': item.get('total_comments'),
                'total_likes': item.get('total_likes'),
                'total_saves': item.get('total_saves'),
                'total_shares': item.get('total_shares'),
                'total_views': item.get('total_views'),
                'uploaded_time': convert_timestamp_to_date_string(item.get('created_at')),
            }
            posts.append(post)

    if filters:
        print("running filter")
        posts = filter_data(posts, filters)

    return posts


def filter_data(data, filters):
    if not filters:
        return data

    hashtags_arr = filters.get("hashtags", "")
    # from_date = filters.get("from", "")
    # to_date = filters.get("to", "")
    result = data
    print("hashtags_arr", hashtags_arr)
    # print("from_date", from_date)
    # print("to_date", to_date)

    # if from_date:
    #     result = list(filter(lambda item: item['created_at'] >= from_date, result))
    # if to_date:
    #     result = list(filter(lambda item: item['created_at'] <= to_date, result))
    if hashtags_arr:
        result = list(filter(lambda item: any(hashtag in item.get('description', '') for hashtag in hashtags_arr), result))

    return result


def tiktok_channel_explore_get_channel_id(search):
    time_sleep = 15
    username = get_tiktok_username(search)
    if username is not None and isinstance(username, str):
        if '/' in username:
            username = username.replace('/', '')
        api_url = baseUrl + f"crawl/tiktok/user-info?username={username}"
        max_retry = 1  # Số lần tối đa gọi lại API khi thất bại
        retry_count = 0

        while retry_count < max_retry:
            # if retry_count == 3 or retry_count == 6:
            #     print(f"{tab_log(tab)}: Sleeping {time_sleep} seconds")
            #     asyncio.sleep(time_sleep)
            #     print(f"{tab_log(tab)}: Sleep done, keeping tracking data")
            response = requests.get(api_url)
            print("api get user id:" + api_url)
            if response.status_code == 200:
                res_data = response.json()
                res_data_data = res_data['data']
                is_success = res_data_data.get('userInfo', None)
                if is_success:
                    user_data = parse_user_data(res_data_data)
                    uid = user_data.get('id')
                    video_count = user_data.get('videoCount')
                    if str(uid).isdigit():
                        return {
                            "uid": uid,
                            "videoCount": video_count,
                        }
                    else:
                        retry_count += 1
                else:
                    retry_count += 1
            else:
                mess_err = tab_log(tab) + f": Retry to get uid user: {username}"
                print(mess_err)
                retry_count += 1

        logger.info(tab_log(tab) + f": Can not find user: {username}")
        # Đã gọi API quá 3 lần mà vẫn thất bại
    return {
        "uid": search,
        "videoCount": 0,
    }


def tiktok_channel_explore_fetch_api(uid, keyword,input_file_id, filters):
    results = []
    if uid:
        max_retry = 1  # Số lần tối đa gọi lại API khi thất bại
        retry_count = 0
        has_more = 1
        maxCursor = ""
        base_api_url = baseUrl + f"crawl/tiktok/user-posts?search={uid}&count=30"

        while has_more and retry_count < max_retry:
            api_url = base_api_url + f"&maxCursor={maxCursor}"
            print(api_url)
            response = requests.get(api_url)
            if response.status_code == 200:
                res_data = response.json()
                data = res_data['data']
                next_page = data.get('next_page', None)
                if next_page:
                    retry_count = 0
                    maxCursor = next_page
                    has_more = data['has_more']

                    data_parse = parse_data(data, keyword,input_file_id, filters)

                    results.append(data_parse)
                else:
                    retry_count += 1
            else:
                print(tab_log(tab) + f": Retry to call api {api_url}")
                retry_count += 1

        if retry_count == max_retry:
            print(tab_log(tab) + f": Call api failed {uid}, maxCursor: {maxCursor}")
            if maxCursor == "":
                results.append(data_failed_default(keyword))
        # Đã gọi API quá 3 lần mà vẫn thất bại
    else:
        results.append(data_failed_default(keyword))

    return results


def channel_explore_get_row(item, platform="tiktok"):
    video_url = item['video']['url'] if 'video' in item and 'url' in item['video'] else None
    video_name = item['name'] if 'name' in item else None
    koc = item['author']['unique_id'] if 'author' in item and 'unique_id' in item['author'] else None
    followers = item['author']['follower_count'] if 'author' in item and 'follower_count' in item[
        'author'] else 0
    views = item['total_views'] if 'total_views' in item else 0
    likes = item['total_likes'] if 'total_likes' in item else 0
    comments = item['total_comments'] if 'total_comments' in item else 0
    shares = item['total_shares'] if 'total_shares' in item else 0
    save = item['total_saves'] if 'total_saves' in item else 0
    uploaded = item['created_at'] if 'created_at' in item else None
    uploaded_format = timestamp_to_date_string(uploaded)
    if platform == "youtube":
        uploaded_format = convert_to_date_format(uploaded)

    return [video_url, video_name, koc, views, likes, comments, shares, save, uploaded_format,
            followers]


def parse_data_kocs(data):
    result = {}

    for video in data:
        if video and len(video):
            author_id = video.get('author', {}).get('unique_id')

            if author_id not in result:
                result[author_id] = {
                    'author': video.get('author'),
                    'total_views': 0,
                    'total_likes': 0,
                    'videos': [],
                    'latest_content': None,
                }

            result[author_id]['total_views'] += video.get('total_views', 0)
            result[author_id]['total_likes'] += video.get('total_likes', 0)

            if result[author_id]['latest_content'] is None or video.get('created_at') > result[author_id][
                'latest_content']:
                result[author_id]['latest_content'] = video.get('created_at')

            result[author_id]['videos'].append(video.get('video'))

    return list(result.values())


async def tiktok_channel_explore_crawl_data(rows, input_file_id, query, url):
    global tab

    # Initialize variables
    max_workers = 8

    # Initialize chunks
    chunks = init_chunks(rows, max_workers)
    try:
        filters = {}
        is_hashtag_mode = get_value_from_query(query, "is_hashtag_mode") == "1"
        print("is_hashtag_mode", is_hashtag_mode)
        if is_hashtag_mode:
            fileRows = await download_and_read_excel_file(url, tab, only_1_row=False)
            filters = get_tiktok_filters(fileRows)
        # Loop through chunks to crawl data
        old_process = 0

        # Get all channel ids
        channel_results = []
        for index, chunk in enumerate(chunks):
            channels = tiktok_channel_explore_get_all_channel_info(chunk, max_workers)
            channel_results = channel_results + channels

        channel_sorted = sorted(channel_results, key=lambda x: x.get("videoCount", 0))

        channel_ids = [item['uid'] for item in channel_sorted]

        chunks = init_chunks_v_array(channel_ids, max_workers)

        for index, chunk in enumerate(chunks):
            try:
                results = tiktok_channel_explore_process_chunk(chunk, filters,input_file_id, max_workers)
                # Append results to the DataFrames
                for result in results:
                    for rows in result:
                        print("rows",rows)
                        import_rows_api(rows,tab)


                # Update process
                now_process = int(((index + 1) / len(chunks)) * 100)
                if now_process - old_process > 0:
                    old_process = now_process
                    update_data = await update_progress_input_file({'id': input_file_id, 'progress': now_process, 'index_processed': index * max_workers})
                    if update_data.get('status', None) == "stop":
                        logger.info(f"{tab_log(tab)}: Stop process")
                        break

            except Exception as e:
                # Handle the error as needed
                print(f"error: {e}")
                pass

        return tab
    except Exception as e:
        logger.info(f"{tab_log(tab)} error: {e}")
    return tab

def get_tiktok_filters(rows):
    hashtag = rows[1][1]

    return {
        "hashtags": hashtag.split(",") if hashtag else [],
    }


def tiktok_channel_explore_get_all_channel_info(chunk,max_workers=3):
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # get channel_ids
        # channel_ids = [(str(tiktok_channel_explore_get_channel_id(search)), search) for search in chunk]
        # call api
        channels = list(executor.map(lambda search: tiktok_channel_explore_get_channel_id(search), chunk))
    return channels

def tiktok_channel_explore_process_chunk(chunk, filters,input_file_id, max_workers=3):
    print(tab_log(tab) + f": Crawling channel chunk: {chunk}")
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # # get channel_ids
        # channel_ids = [(str(tiktok_channel_explore_get_channel_id(search)), search) for search in chunk]
        # call api
        results = list(executor.map(lambda x: tiktok_channel_explore_fetch_api(x, x,input_file_id, filters), chunk))
    return results


def youtube_channel_explore_process_chunk(chunk, max_workers=3):
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # get channel_ids
        channel_urls = [(str(get_youtube_channel_url(search)), search) for search in chunk]
        # call api
        results = list(executor.map(lambda url: youtube_channel_explore_fetch_api(str(url[0])), channel_urls))
    return results


def instagram_channel_explore_process_chunk(chunk, max_workers=3):
    print(tab_log(tab) + f": Crawling chunk: {chunk}")
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        # get channel_ids
        users = [((instagram_channel_explore_get_user_info(search)), search) for search in chunk]
        # call api
        results = list(executor.map(lambda x: instagram_channel_explore_fetch_api(x[0]), users))
    return results


def get_username_from_instagram_urls(url):
    username_regex = r'instagram\.com\/([A-Za-z0-9._]+)'
    match = re.search(username_regex, url)
    if match and len(match.groups()) > 0:
        return match.group(1)
    else:
        return url


def instagram_parse_data_user(data, koc):
    user = {
        "koc": koc,
        "ads_page_name": data.get("ads_page_name"),
        "full_name": data.get("full_name"),
        "username": data.get("username"),
        "unique_id": data.get("username"),
        "follower_count": data.get("follower_count"),
        "media_count": data.get("media_count"),
        "avatar_url": data.get("hd_profile_pic_versions")[0].get("url") if data.get(
            "hd_profile_pic_versions") else None,
    }
    return user


def instagram_channel_explore_get_user_info(search):
    username = get_username_from_instagram_urls(search)
    api_url = baseUrl + f"crawl/instagram/user-info?user_name={username}"
    max_retry = 5  # Số lần tối đa gọi lại API khi thất bại
    retry_count = 0

    while retry_count < max_retry:
        print(api_url)
        response = requests.get(api_url)
        try:
            if response.status_code == 200:
                res_data = response.json()
                res_data_data = res_data['data']
                try:
                    is_success = res_data_data.get('username', None)
                except Exception as e:
                    is_success = False
                if is_success:
                    user_data = instagram_parse_data_user(res_data['data'], search)
                    if user_data:
                        return user_data
                    else:
                        retry_count += 1
                else:
                    retry_count += 1
            else:
                mess_err = tab_log(tab) + f": Retry to get uid user: {username}"
                print(mess_err)
                retry_count += 1
        except Exception as e:
            retry_count += 1

    logger.info(tab_log(tab) + f": Can not find user: {username}")
    # Đã gọi API quá 3 lần mà vẫn thất bại
    return search


def instagram_parse_data(data, follower_count):
    posts = []
    if data and 'edges' in data:
        for edge in data['edges']:
            try:
                item = edge.get('node', {})
                post = {
                    'id': item.get('shortcode', None),
                    'name': item['edge_media_to_caption']['edges'][0]['node'][
                        'text'] if 'edge_media_to_caption' in item and 'edges' in item['edge_media_to_caption'] and
                                   item['edge_media_to_caption']['edges'] else None,
                    'video': {
                        'thumb': item.get('thumbnail_resources')[0]['src'] if 'thumbnail_resources' in item and item[
                            'thumbnail_resources'] else None,
                        'url': 'https://instagram.com/p/' + item.get('shortcode',
                                                                     None) if 'shortcode' in item else None,
                    },
                    'author': {
                        'id': item.get('owner', {}).get('id', None),
                        'unique_id': item.get('owner', {}).get('username', None),
                        'shortcode': item.get('shortcode', None),
                        'follower_count': follower_count
                    },
                    'total_comments': item['edge_media_to_comment']['count'] if 'edge_media_to_comment' in item else 0,
                    'total_likes': item['edge_media_preview_like']['count'] if 'edge_media_preview_like' in item and
                                                                               item['edge_media_preview_like'][
                                                                                   'count'] != -1 else None,
                    'created_at': item.get('taken_at_timestamp', None),
                }
                posts.append(post)

            except Exception as e:
                logger.info(f"{tab_log(tab)}: cannot append new edge | err: {e}")

    return posts


def instagram_channel_explore_fetch_api(user):
    uid = user.get('unique_id', None)
    follower_count = user.get('follower_count', 0)
    results = []

    if uid:
        max_retry = 3  # Số lần tối đa gọi lại API khi thất bại
        retry_count = 0
        has_more = 1
        end_cursor = ""
        base_api_url = baseUrl + f"crawl/instagram/user-posts?user_name={uid}"

        while has_more and retry_count < max_retry:
            api_url = base_api_url
            if end_cursor:
                api_url = base_api_url + f"&end_cursor={end_cursor}"

            print(api_url)
            response = requests.get(api_url)
            if response.status_code == 200:
                try:
                    res_data = response.json()
                    data = res_data['data']
                    page = data.get("page_info", {})
                    next_page = page.get('end_cursor', None)
                    has_next_page = page.get('has_next_page', None)
                    print("next_page", next_page)
                    print("has_next_page", has_next_page)
                    if page:
                        retry_count = 0
                        end_cursor = next_page
                        has_more = has_next_page
                        new_data = instagram_parse_data(data, follower_count)
                        for item in new_data:
                            row_posts = channel_explore_get_row(item)
                            cleaned_row = [remove_illegal_characters(cell_value) for cell_value in row_posts]
                            results.append(cleaned_row)
                    else:
                        retry_count += 1

                except Exception as e:
                    retry_count += 1
                    print(f"{tab_log(tab)}: Cannot handle result data | err: {e}")
            else:
                print(tab_log(tab) + f": Retry to call api {tab}")
                retry_count += 1

    return results
    # return instagram_failed_default(user.get('keyword',None))


async def instagram_channel_explore_crawl_data(rows, input_file_id, tab="instagram-channel-explore"):
    # Initialize variables
    file_path = set_file_path(tab)
    max_workers = 2
    sheet_koc_name = "Unique KOCs"

    wb = openpyxl.Workbook()
    # Select the active sheet (you can modify this depending on your Excel structure)
    sheet_posts = wb.create_sheet(title="Posts")
    sheet_kocs = wb.create_sheet(title=sheet_koc_name)

    # Initialize header rows
    sheet_posts.append(channel_explore_header_row_post)
    sheet_kocs.append(channel_explore_header_row_kocs)

    # Remove default sheet
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    # Save the workbook once before processing chunks
    wb.save(file_path)

    # Initialize chunks
    chunks = init_chunks(rows, max_workers)

    try:
        # Loop through chunks to crawl data
        old_process = 0
        for index, chunk in enumerate(chunks):
            try:
                results = instagram_channel_explore_process_chunk(chunk, max_workers)

                # update progress
                now_process = int(((index + 1) / len(chunks)) * 100)
                if now_process - old_process > 0:
                    old_process = now_process
                    update_data = await update_progress_input_file({'id': input_file_id, 'progress': now_process})
                    if update_data.get('status', None) == "stop":
                        logger.info(f"{tab_log(tab)}: Stop process")
                        break

                # Append results to the sheet
                for result in results:
                    for row in result:
                        sheet_posts.append(row)

            except Exception as e:
                wb.save(file_path)
                logger.info(f"{tab_log(tab)}: chunk: {chunk} | error: {e}")

        # Save the workbook after processing all chunks
        wb.save(file_path)

        # Close the workbook before processing KOC data
        wb.close()

        # Process the KOC data separately
        channel_handle_unique_koc_sheet_data(file_path, sheet_koc_name)

        return file_path

    except Exception as e:
        # In case of any exception, save the workbook and close it before re-raising the exception
        wb.save(file_path)
        wb.close()
        raise e


async def youtube_channel_explore_crawl_data(rows, input_file_id, tab="youtube-channel-explore"):
    # Initialize variables
    file_path = set_file_path(tab)
    max_workers = 1
    sheet_koc_name = "Unique KOCs"

    wb = openpyxl.Workbook()
    # Select the active sheet (you can modify this depending on your Excel structure)
    sheet_posts = wb.create_sheet(title="Posts")
    sheet_kocs = wb.create_sheet(title=sheet_koc_name)

    # Initialize header rows
    sheet_posts.append(channel_explore_header_row_post)
    sheet_kocs.append(channel_explore_header_row_kocs)

    # Remove default sheet
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)

    # Save the workbook once before processing chunks
    wb.save(file_path)

    # Initialize chunks
    chunks = init_chunks(rows, max_workers)

    try:
        # Loop through chunks to crawl data
        old_process = 0
        for index, chunk in enumerate(chunks):
            try:
                results = youtube_channel_explore_process_chunk(chunk, max_workers)
                # update process
                now_process = int(((index + 1) / len(chunks)) * 100)
                if now_process - old_process > 0:
                    old_process = now_process
                    update_data = await update_progress_input_file({'id': input_file_id, 'progress': now_process})
                    if update_data.get('status', None) == "stop":
                        logger.info(f"{tab_log(tab)}: Stop process")
                        break

                # Append results to the sheet
                for result in results:
                    for row in result:
                        print("row", row)
                        try:
                            sheet_posts.append(row)
                        except Exception as e:
                            logger.info(f"{tab_log(tab)}: row: {row} | error: {e}")

            except Exception as e:
                wb.save(file_path)
                logger.info(f"{tab_log(tab)}: chunk: {chunk} | error: {e}")

        # Save the workbook after processing all chunks
        wb.save(file_path)

        # Close the workbook before processing KOC data
        wb.close()

        # Process the KOC data separately
        channel_handle_unique_koc_sheet_data(file_path, sheet_koc_name)

        return file_path

    except Exception as e:
        # In case of any exception, save the workbook and close it before re-raising the exception
        wb.save(file_path)
        wb.close()
        raise e


def youtube_channel_explore_fetch_api(url):
    max_retry = 3  # Số lần tối đa gọi lại API khi thất bại
    retry_count = 0
    page = 1
    base_api_url = baseUrl + f"crawl/youtube/channel-posts?channelId={url}&limit=30"
    continuously_fail = 0
    max_continuously_fail = 3

    results = []
    while retry_count < max_retry and continuously_fail < max_continuously_fail:
        api_url = base_api_url + f"&page={page}"
        print(api_url)
        response = requests.get(api_url)
        if response.status_code == 200:
            res_data = response.json()
            data = res_data['data']
            if len(data):
                continuously_fail = 0
                page += 1
                retry_count = 0
                try:
                    new_data = youtube_parse_data(data, get_youtube_username_from_url(url))
                    for item in new_data:
                        row_posts = channel_explore_get_row(item, "youtube")
                        cleaned_row = [remove_illegal_characters(cell_value) for cell_value in row_posts]
                        results.append(cleaned_row)
                except Exception as e:
                    print(f"Append failed {e}:", data)
            else:
                print(f"{tab_log(tab)}: continuously_fail + 1")
                continuously_fail += 1

        else:
            print(tab_log(tab) + f": Retry to call api {api_url}")
            retry_count += 1

    # Đã gọi API quá 3 lần mà vẫn thất bại
    if page == 1 and retry_count == max_retry:
        results.append(data_failed_default(url))
    return results


def youtube_parse_data(data, username):
    posts = []
    for item in data:
        try:
            post = {
                "id": item.get("id"),
                "name": item.get("title"),
                "video": {
                    "thumb": item.get("thumbnail"),
                    "url": f"https://www.youtube.com/watch?v={item.get('id')}",
                },
                "author": {
                    "name": item.get("channel", {}).get("name"),
                    "avatar_url": item.get("channel", {}).get("avatar_thumb", {}).get("url_list", [False])[0],
                    "uid": item.get("channel", {}).get("id"),
                    "unique_id": username,
                    "follower_count": item.get("channel", {}).get("channel_follower_count"),
                },
                "total_comments": item.get("commentCount", 0),
                "total_likes": item.get("likeCount", 0),
                "total_views": item.get("viewCount", 0),
                "created_at": item.get("uploadDate"),
            }
            posts.append(post)
        except e:
            logger.info(f"{tab_log(tab)}: Cannot append {item}")

    return posts


def is_youtube_url(url):
    return is_valid_http_url(url) and 'youtube.com' in url


def get_youtube_channel_url(keyword):
    un = keyword if is_youtube_url(keyword) else f"https://youtube.com/{keyword}"
    return un


def get_youtube_username_from_url(url):
    # Kiểm tra URL có hợp lệ và bắt đầu bằng "https://www.youtube.com/@"
    if url.startswith("https://www.youtube.com/@"):
        # Tìm vị trí của ký tự "-" trong URL
        index = url.rfind("-")

        # Trích xuất tên người dùng từ URL
        if index != -1:
            username = url[len("https://www.youtube.com/@"):index]
        else:
            username = url[len("https://www.youtube.com/@"):]

        return f"@{username}"
    else:
        return url
